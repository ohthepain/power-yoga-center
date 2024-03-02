import sys, os
import importlib
import thrift
from thrift.transport import TTransport
from thrift.protocol import TJSONProtocol
from thrift.Thrift import TType
from openpyxl import Workbook
from openpyxl import load_workbook

sys.path.append("../Thrift/gen-py")
sys.path.append("../Thrift/gen-py/Yoga")
ThriftModule = importlib.import_module('Yoga.Config.ttypes')
ThriftConstants = importlib.import_module('Yoga.Config.constants')
SharedModule = importlib.import_module('shared.ttypes')
SharedConstants = importlib.import_module('shared.constants')
Data = ThriftModule.Data()

typeRecord = {}
sheetInfo = None

def stringToName(s):
	name = s.split("--", 1)[0].strip().replace(" ", "")
	return name[0].lower() + name[1:]

class SheetColumnEntry:
	title = None
	attributeName = None
	columnNum = 0
	def __init__(self, cnum, topcell):
		self.title = topcell
		self.attributeName = stringToName(topcell)
		self.columnNum = cnum
		self.dump()
	def dump(self):
		print("SheetColumnEntry for column %d has attributeName %s (%s)" % (self.columnNum, self.attributeName, self.title))

class SheetInfo:
	title = None
	tablename = None
	columnInfo = None
	def __init__(self, sheet):
		self.title = sheet.title
		self.tablename = stringToName(self.title)
		self.columnInfo = {}
		print("==>" + sheet.title + " - " + self.tablename)
		for c in range(1,101):
			topcell = sheet.cell(row=1, column=c).value
			if 	topcell == None:
				break
			print("--------------------------%d %s" % (c, topcell))
			entry = SheetColumnEntry(c, topcell)
			self.columnInfo[c] = entry
			entry.dump()
	def dump(self):
		print("SheetInfo (%s) : %s" % (self.title, self.tablename))
		for c in range(1,len(self.columnInfo)):
			columnInfo = self.columnInfo[c]
			if columnInfo != None:
				print("	column %d: column %d attribute name %s" % (c, columnInfo.columnNum, columnInfo.attributeName))

class TypeEntry:
	type = None
	name = None
	mapIndexType = None
	mapValueType = None
	listValueType = None
	def __init__(self, spec):
		if spec != None:
			self.columnNum = spec[0]
			self.type = spec[1]
			self.name = spec[2]
			if self.type == TType.MAP:
				self.mapIndexType = spec[3][0]
				self.mapValueType = spec[3][3][0]
				# this constructs the object
				new_object = self.mapValueType()
			elif self.type == TType.LIST:
				self.listValueType = spec[3][1][0]
				# this constructs the object
				new_object = self.listValueType()

def makeTableClassMap(Data):
	for spec in Data.thrift_spec:
		typeEntry = TypeEntry(spec)
		if typeEntry.mapIndexType != None:
			print("Set type entry for map " + typeEntry.name)
			print(typeEntry.mapValueType)
			typeRecord[typeEntry.name] = typeEntry
		elif typeEntry.listValueType != None:
			print("Set type entry for list " + typeEntry.name)
			print(typeEntry.listValueType)
			typeRecord[typeEntry.name] = typeEntry

def countActiveColumns(sheet):
	column = 1
	while sheet.cell(row=1, column=column).value != None:
		column += 1
	return column - 1

def parseSheet(sheet):
	sheetInfo = SheetInfo(sheet)
	print("++++++++++++++++++++++++++++++")
	sheetInfo.dump()
	print("++++++++++++++++++++++++++++++")
	tablename = stringToName(sheet.title)
	typeEntry = typeRecord[tablename]
	if typeEntry != None:
		print("parseSheet " + sheet.title + " with tablename " + tablename)
		table = getattr(Data, tablename)
		if table == None:
			if typeEntry.mapValueType != None:
				setattr(Data, tablename, {})
			elif typeEntry.listValueType != None:
				setattr(Data, tablename, [])
			table = getattr(Data, tablename)
		# traverse rows
		row = 2
		while sheet.cell(row=row, column=1).value != None:
			if typeEntry.mapValueType != None:
				new_object = typeEntry.mapValueType()
				# populate object
				for column in range(1, countActiveColumns(sheet) + 1):
					value = sheet.cell(row=row, column=column).value
					if value == None:
						value = ""
					# get member for column
					info = sheetInfo.columnInfo[column]
					#print("set attribute %s column %d to value %s" % (info.attributeName, column, value))
					setattr(new_object, info.attributeName, value)
				indexAttributeName = sheetInfo.columnInfo[1].attributeName
				#print("index attribute name is " + indexAttributeName)
				objectIndex = getattr(new_object, indexAttributeName)
				#print("index attribute value is " + objectIndex)
				table[objectIndex] = new_object
			elif typeEntry.listValueType != None:
				new_object = typeEntry.listValueType()
				# populate object
				for column in range(1, countActiveColumns(sheet) + 1):
					value = sheet.cell(row=row, column=column).value
					if value == None:
						value = ""
					# get member for column
					info = sheetInfo.columnInfo[column]
					#print("set attribute %s column %d to value %s" % (info.attributeName, column, value))
					setattr(new_object, info.attributeName, value)
				table.append(new_object)
			row += 1

def parseWorkbook(workbook):
	print("parseWorkbook ")
	for sheet in workbook:
		parseSheet(sheet)

def parseFile(path):
	workbook = load_workbook(path, read_only = True, data_only = True)
	parseWorkbook(workbook)

members = [attr for attr in dir(Data) if not callable(getattr(Data, attr)) and not attr.startswith("__")]
print "Data members: "
print members

makeTableClassMap(Data)

for root, dirs, files in os.walk("."):
	for file_ in files:
		if file_[:2] != "~$" and file_.lower().endswith(".xlsx"):
			parseFile(os.path.join(root, file_))


print("Data:")
print(Data)


#ThriftProtocol = getattr(importlib.import_module('thrift.protocol.%s' % args.tprotocol), args.tprotocol)
#ThriftModule = importlib.import_module('Yoga.Config.ttypes' % args.namespace)
#ThriftConstants = importlib.import_module('%s.constants' % args.namespace)

#ThriftType = getattr(ThriftModule, 'Data')

transport = TTransport.TMemoryBuffer()
ThriftProtocol = getattr(importlib.import_module('thrift.protocol.TJSONProtocol'), 'TJSONProtocol')
protocol = ThriftProtocol(transport)
Data.schemaVersionId = 1;
Data.write(protocol)

buf = transport.getvalue()
with open("../../config.bin", 'wb') as f:
	f.write(buf)

buf = None

with open("../../config.bin", 'rb') as f:
	buf = f.read()

transport = TTransport.TMemoryBuffer(buf)
protocol = ThriftProtocol(transport)
Data = None
Data = ThriftModule.Data()
#transport.setvalue(buf)
Data.read(protocol)

print("Data2:")
print(Data)
print("eosData2:")

